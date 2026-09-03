from __future__ import annotations

import csv
import io
import logging
import math
import re
import uuid
from datetime import datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.dependencies import CurrentUser
from app.core.exceptions import (
    ConflictError,
    NotFoundError,
    PermissionDeniedError,
    ValidationError,
)
from app.core.schemas import PaginatedResponse
from app.core.storage import get_object_storage
from app.modules.agents.models import Agent, AgentStatus
from app.modules.billing.entitlements import EntitlementService
from app.modules.outbound_campaigns.models import (
    CampaignStatus,
    CampaignType,
    DoNotCallEntry,
    OutboundAttempt,
    OutboundCampaign,
    OutboundRecipient,
    RecipientStatus,
)
from app.modules.outbound_campaigns.schemas import (
    AudioGenerateRequest,
    AudioPlaybackResponse,
    AudioResponse,
    CampaignCreate,
    CampaignResponse,
    CampaignScheduleRequest,
    CampaignTestCallRequest,
    CampaignUpdate,
    CampaignValidationResponse,
    DoNotCallCreate,
    ImportErrorRow,
    ImportResponse,
    OutboundEventRequest,
    RecipientResponse,
    SingleOutboundCallRequest,
)
from app.modules.outbound_campaigns.tts import CampaignTTS
from app.modules.phone_connections.providers import AsteriskProvisionerClient
from app.modules.phone_numbers.models import ConnectionStatus, PhoneNumber

logger = logging.getLogger(__name__)

E164 = re.compile(r"^\+[1-9]\d{7,14}$")
TERMINAL = {
    RecipientStatus.COMPLETED,
    RecipientStatus.CANCELLED,
    RecipientStatus.DO_NOT_CALL,
}


class OutboundCampaignService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self.entitlements = EntitlementService(db)

    @staticmethod
    def _company_id(user: CurrentUser) -> uuid.UUID:
        if not user.company_id:
            raise PermissionDeniedError("No company context")
        return uuid.UUID(user.company_id)

    async def _get(self, campaign_id: uuid.UUID, user: CurrentUser) -> OutboundCampaign:
        campaign = await self.db.scalar(
            select(OutboundCampaign).where(
                OutboundCampaign.id == campaign_id,
                OutboundCampaign.company_id == self._company_id(user),
            )
        )
        if not campaign:
            raise NotFoundError("Outbound campaign not found")
        return campaign

    async def _validate_resources(
        self, data: CampaignCreate, company_id: uuid.UUID
    ) -> None:
        phone = await self.db.scalar(
            select(PhoneNumber).where(
                PhoneNumber.id == data.phone_number_id,
                PhoneNumber.company_id == company_id,
                PhoneNumber.is_enabled.is_(True),
                PhoneNumber.connection_status == ConnectionStatus.CONNECTED,
            )
        )
        if not phone or not phone.connection_id:
            raise ValidationError("An enabled, connected phone number is required")
        if data.agent_id:
            agent = await self.db.scalar(
                select(Agent).where(
                    Agent.id == data.agent_id,
                    Agent.company_id == company_id,
                    Agent.status == AgentStatus.ACTIVE,
                )
            )
            if not agent:
                raise ValidationError("An active tenant agent is required")
        try:
            ZoneInfo(data.timezone)
        except ZoneInfoNotFoundError as exc:
            raise ValidationError("Unknown campaign timezone") from exc

    async def _response(self, campaign: OutboundCampaign) -> CampaignResponse:
        counts = dict(
            (
                await self.db.execute(
                    select(OutboundRecipient.status, func.count(OutboundRecipient.id))
                    .where(OutboundRecipient.campaign_id == campaign.id)
                    .group_by(OutboundRecipient.status)
                )
            ).all()
        )
        total = sum(counts.values())
        completed = counts.get(RecipientStatus.COMPLETED, 0)
        failed = sum(
            counts.get(status, 0)
            for status in (
                RecipientStatus.FAILED,
                RecipientStatus.BUSY,
                RecipientStatus.NO_ANSWER,
            )
        )
        return CampaignResponse(
            **{
                key: getattr(campaign, key)
                for key in (
                    "id",
                    "company_id",
                    "name",
                    "campaign_type",
                    "status",
                    "agent_id",
                    "phone_number_id",
                    "message_text",
                    "voice",
                    "language",
                    "scheduled_at",
                    "timezone",
                    "calling_window_start",
                    "calling_window_end",
                    "max_concurrency",
                    "max_attempts",
                    "retry_delay_minutes",
                    "ring_timeout_seconds",
                    "keypad_actions",
                    "settings",
                    "created_at",
                    "updated_at",
                )
            },
            audio_ready=bool(campaign.audio_media_id),
            total_recipients=total,
            completed_recipients=completed,
            failed_recipients=failed,
        )

    async def create(self, data: CampaignCreate, user: CurrentUser) -> CampaignResponse:
        company_id = self._company_id(user)
        await self.entitlements.require_active_subscription(company_id)
        await self._validate_resources(data, company_id)
        concurrency = min(
            data.max_concurrency, settings.OUTBOUND_MAX_CONCURRENCY_PER_COMPANY
        )
        campaign = OutboundCampaign(
            **data.model_dump(exclude={"max_concurrency"}),
            max_concurrency=concurrency,
            company_id=company_id,
            created_by=uuid.UUID(user.user_id),
        )
        self.db.add(campaign)
        await self.db.commit()
        await self.db.refresh(campaign)
        return await self._response(campaign)

    async def list(
        self, user: CurrentUser, page: int, page_size: int
    ) -> PaginatedResponse[CampaignResponse]:
        company_id = self._company_id(user)
        query = select(OutboundCampaign).where(
            OutboundCampaign.company_id == company_id
        )
        total = int(
            await self.db.scalar(select(func.count()).select_from(query.subquery()))
            or 0
        )
        items = list(
            await self.db.scalars(
                query.order_by(OutboundCampaign.created_at.desc())
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        )
        return PaginatedResponse(
            items=[await self._response(item) for item in items],
            total=total,
            page=page,
            page_size=page_size,
            pages=math.ceil(total / page_size) if total else 0,
        )

    async def get(self, campaign_id: uuid.UUID, user: CurrentUser) -> CampaignResponse:
        return await self._response(await self._get(campaign_id, user))

    async def update(
        self, campaign_id: uuid.UUID, data: CampaignUpdate, user: CurrentUser
    ) -> CampaignResponse:
        campaign = await self._get(campaign_id, user)
        if campaign.status not in (
            CampaignStatus.DRAFT,
            CampaignStatus.READY,
            CampaignStatus.PAUSED,
        ):
            raise ConflictError("Only draft, ready, or paused campaigns can be edited")
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(campaign, key, value)
        campaign.audio_media_id = (
            None
            if {"message_text", "voice"} & data.model_fields_set
            else campaign.audio_media_id
        )
        await self.db.commit()
        await self.db.refresh(campaign)
        return await self._response(campaign)

    @staticmethod
    def _read_rows(filename: str, content: bytes) -> list[dict[str, str]]:
        suffix = filename.lower().rsplit(".", 1)[-1]
        if suffix == "csv":
            text = content.decode("utf-8-sig")
            return [
                {str(k).strip().lower(): (v or "").strip() for k, v in row.items()}
                for row in csv.DictReader(io.StringIO(text))
            ]
        if suffix == "xlsx":
            workbook = load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip().lower() for value in next(values, [])]
            return [
                {
                    headers[index]: str(value).strip() if value is not None else ""
                    for index, value in enumerate(row)
                    if index < len(headers)
                }
                for row in values
            ]
        raise ValidationError("Only .csv and .xlsx files are supported")

    async def import_contacts(
        self,
        campaign_id: uuid.UUID,
        filename: str,
        content: bytes,
        user: CurrentUser,
        replace: bool,
    ) -> ImportResponse:
        campaign = await self._get(campaign_id, user)
        if campaign.status not in (CampaignStatus.DRAFT, CampaignStatus.READY):
            raise ConflictError(
                "Contacts can only be imported into draft or ready campaigns"
            )
        rows = self._read_rows(filename, content)
        if len(rows) > settings.OUTBOUND_MAX_IMPORT_ROWS:
            raise ValidationError(
                f"Import is limited to {settings.OUTBOUND_MAX_IMPORT_ROWS} rows"
            )
        if replace:
            await self.db.execute(
                delete(OutboundRecipient).where(
                    OutboundRecipient.campaign_id == campaign.id
                )
            )
        existing = set(
            await self.db.scalars(
                select(OutboundRecipient.phone_number).where(
                    OutboundRecipient.campaign_id == campaign.id
                )
            )
        )
        dnc = set(
            await self.db.scalars(
                select(DoNotCallEntry.phone_number).where(
                    DoNotCallEntry.company_id == campaign.company_id
                )
            )
        )
        imported = duplicates = rejected = 0
        errors: list[ImportErrorRow] = []
        known = set(existing)
        standard = {
            "phone_number",
            "first_name",
            "last_name",
            "language",
            "timezone",
            "external_id",
            "consent_at",
            "do_not_call",
        }
        for row_number, row in enumerate(rows, start=2):
            phone = row.get("phone_number", "").replace(" ", "")
            if not E164.fullmatch(phone):
                rejected += 1
                errors.append(
                    ImportErrorRow(
                        row=row_number,
                        phone_number=phone or None,
                        error="phone_number must be E.164",
                    )
                )
                continue
            if phone in known:
                duplicates += 1
                continue
            known.add(phone)
            consent_at = None
            if row.get("consent_at"):
                try:
                    consent_at = datetime.fromisoformat(
                        row["consent_at"].replace("Z", "+00:00")
                    )
                except ValueError:
                    rejected += 1
                    errors.append(
                        ImportErrorRow(
                            row=row_number,
                            phone_number=phone,
                            error="consent_at must be ISO-8601",
                        )
                    )
                    continue
            blocked = phone in dnc or row.get("do_not_call", "").lower() in {
                "1",
                "true",
                "yes",
            }
            recipient = OutboundRecipient(
                campaign_id=campaign.id,
                company_id=campaign.company_id,
                phone_number=phone,
                first_name=row.get("first_name") or None,
                last_name=row.get("last_name") or None,
                language=row.get("language") or None,
                timezone=row.get("timezone") or None,
                external_id=row.get("external_id") or None,
                consent_at=consent_at,
                custom_fields={
                    key: value
                    for key, value in row.items()
                    if key not in standard and value
                }
                or None,
                status=RecipientStatus.DO_NOT_CALL
                if blocked
                else RecipientStatus.PENDING,
            )
            self.db.add(recipient)
            imported += 1
        campaign.status = (
            CampaignStatus.READY if imported or existing else CampaignStatus.DRAFT
        )
        try:
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise ConflictError("The import contains duplicate recipients") from exc
        return ImportResponse(
            imported=imported,
            duplicates=duplicates,
            rejected=rejected,
            errors=errors[:200],
        )

    async def recipients(
        self, campaign_id: uuid.UUID, user: CurrentUser, page: int, page_size: int
    ) -> PaginatedResponse[RecipientResponse]:
        campaign = await self._get(campaign_id, user)
        query = select(OutboundRecipient).where(
            OutboundRecipient.campaign_id == campaign.id
        )
        total = int(
            await self.db.scalar(select(func.count()).select_from(query.subquery()))
            or 0
        )
        rows = list(
            await self.db.scalars(
                query.order_by(OutboundRecipient.created_at)
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        )
        return PaginatedResponse(
            items=[RecipientResponse.model_validate(row) for row in rows],
            total=total,
            page=page,
            page_size=page_size,
            pages=math.ceil(total / page_size) if total else 0,
        )

    async def validate(
        self, campaign_id: uuid.UUID, user: CurrentUser
    ) -> CampaignValidationResponse:
        campaign = await self._get(campaign_id, user)
        counts = dict(
            (
                await self.db.execute(
                    select(OutboundRecipient.status, func.count())
                    .where(OutboundRecipient.campaign_id == campaign.id)
                    .group_by(OutboundRecipient.status)
                )
            ).all()
        )
        total = sum(counts.values())
        blocked = counts.get(RecipientStatus.DO_NOT_CALL, 0)
        errors = []
        if total == 0:
            errors.append("Import at least one recipient")
        if (
            campaign.campaign_type != CampaignType.AI_CONVERSATION
            and not campaign.audio_media_id
        ):
            errors.append("Generate and approve the broadcast audio")
        callable_count = total - blocked
        return CampaignValidationResponse(
            valid=not errors and callable_count > 0,
            total=total,
            callable=callable_count,
            invalid=0,
            do_not_call=blocked,
            errors=errors,
        )

    async def generate_audio(
        self,
        campaign_id: uuid.UUID,
        user: CurrentUser,
        data: AudioGenerateRequest | None = None,
    ) -> AudioResponse:
        campaign = await self._get(campaign_id, user)
        if data:
            campaign.message_text = data.message_text
            if data.voice:
                campaign.voice = data.voice
            campaign.audio_media_id = None
            campaign.audio_storage_key = None
        if (
            campaign.campaign_type == CampaignType.AI_CONVERSATION
            or not campaign.message_text
        ):
            raise ValidationError("This campaign does not have broadcast audio")
        media_id, wav = await CampaignTTS().generate_wav(
            text=campaign.message_text, voice=campaign.voice
        )
        key = f"outbound/{campaign.company_id}/{media_id}.wav"
        await get_object_storage().upload(
            io.BytesIO(wav), key=key, content_type="audio/wav"
        )
        await AsteriskProvisionerClient().upload_outbound_media(media_id, wav)
        campaign.audio_storage_key = key
        campaign.audio_media_id = media_id
        await self.db.commit()
        logger.info(
            "Generated outbound campaign audio campaign_id=%s media_id=%s voice=%s",
            campaign.id,
            media_id,
            campaign.voice,
        )
        return AudioResponse(audio_ready=True, storage_key=key, media_id=media_id)

    async def get_audio_url(
        self, campaign_id: uuid.UUID, user: CurrentUser
    ) -> AudioPlaybackResponse:
        campaign = await self._get(campaign_id, user)
        if not campaign.audio_storage_key or not campaign.audio_media_id:
            raise NotFoundError("Campaign audio has not been generated")
        expires_in = settings.RECORDING_PRESIGNED_URL_EXPIRE_SECONDS
        url = await get_object_storage().presigned_download_url(
            key=campaign.audio_storage_key,
            expires_in=expires_in,
        )
        return AudioPlaybackResponse(
            url=url,
            expires_in_seconds=expires_in,
            media_id=campaign.audio_media_id,
        )

    async def schedule(
        self, campaign_id: uuid.UUID, data: CampaignScheduleRequest, user: CurrentUser
    ) -> CampaignResponse:
        campaign = await self._get(campaign_id, user)
        await self.entitlements.require_minutes_available(campaign.company_id, lock=True)
        validation = await self.validate(campaign_id, user)
        if not validation.valid:
            raise ValidationError(
                "Campaign is not ready", {"errors": validation.errors}
            )
        campaign.scheduled_at = data.scheduled_at.astimezone(timezone.utc)
        campaign.status = CampaignStatus.SCHEDULED
        await self.db.commit()
        return await self._response(campaign)

    async def start(
        self, campaign_id: uuid.UUID, user: CurrentUser
    ) -> CampaignResponse:
        campaign = await self._get(campaign_id, user)
        await self.entitlements.require_minutes_available(campaign.company_id, lock=True)
        validation = await self.validate(campaign_id, user)
        if not validation.valid:
            raise ValidationError(
                "Campaign is not ready", {"errors": validation.errors}
            )
        campaign.status = CampaignStatus.RUNNING
        campaign.scheduled_at = datetime.now(timezone.utc)
        await self.db.commit()
        from app.workers.outbound_tasks import dispatch_campaign

        dispatch_campaign.delay(str(campaign.id))
        return await self._response(campaign)

    async def set_status(
        self, campaign_id: uuid.UUID, status: CampaignStatus, user: CurrentUser
    ) -> CampaignResponse:
        campaign = await self._get(campaign_id, user)
        allowed = {
            CampaignStatus.PAUSED: {CampaignStatus.RUNNING, CampaignStatus.SCHEDULED},
            CampaignStatus.RUNNING: {CampaignStatus.PAUSED},
            CampaignStatus.CANCELLED: {
                CampaignStatus.DRAFT,
                CampaignStatus.READY,
                CampaignStatus.SCHEDULED,
                CampaignStatus.RUNNING,
                CampaignStatus.PAUSED,
            },
        }
        if campaign.status not in allowed[status]:
            raise ConflictError(
                f"Cannot change {campaign.status.value} campaign to {status.value}"
            )
        if status == CampaignStatus.RUNNING:
            await self.entitlements.require_minutes_available(
                campaign.company_id, lock=True
            )
        campaign.status = status
        if status == CampaignStatus.CANCELLED:
            await self.db.execute(
                OutboundRecipient.__table__.update()
                .where(
                    OutboundRecipient.campaign_id == campaign.id,
                    OutboundRecipient.status.in_(
                        [RecipientStatus.PENDING, RecipientStatus.QUEUED]
                    ),
                )
                .values(status=RecipientStatus.CANCELLED)
            )
        await self.db.commit()
        if status == CampaignStatus.RUNNING:
            from app.workers.outbound_tasks import dispatch_campaign

            dispatch_campaign.delay(str(campaign.id))
        return await self._response(campaign)

    async def export_results(self, campaign_id: uuid.UUID, user: CurrentUser) -> bytes:
        campaign = await self._get(campaign_id, user)
        rows = list(
            await self.db.scalars(
                select(OutboundRecipient)
                .where(OutboundRecipient.campaign_id == campaign.id)
                .order_by(OutboundRecipient.created_at)
            )
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(
            [
                "phone_number",
                "first_name",
                "last_name",
                "external_id",
                "status",
                "attempts",
                "last_call_id",
                "last_error",
            ]
        )
        for row in rows:
            sheet.append(
                [
                    row.phone_number,
                    row.first_name,
                    row.last_name,
                    row.external_id,
                    row.status.value,
                    row.attempts_count,
                    str(row.last_call_id or ""),
                    row.last_error,
                ]
            )
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def add_dnc(self, data: DoNotCallCreate, user: CurrentUser) -> DoNotCallEntry:
        if not E164.fullmatch(data.phone_number):
            raise ValidationError("phone_number must be E.164")
        entry = DoNotCallEntry(company_id=self._company_id(user), **data.model_dump())
        self.db.add(entry)
        try:
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise ConflictError(
                "Phone number is already on the do-not-call list"
            ) from exc
        await self.db.refresh(entry)
        return entry

    async def list_dnc(self, user: CurrentUser) -> list[DoNotCallEntry]:
        return list(
            await self.db.scalars(
                select(DoNotCallEntry)
                .where(DoNotCallEntry.company_id == self._company_id(user))
                .order_by(DoNotCallEntry.created_at.desc())
            )
        )

    async def delete_dnc(self, entry_id: uuid.UUID, user: CurrentUser) -> None:
        entry = await self.db.scalar(
            select(DoNotCallEntry).where(
                DoNotCallEntry.id == entry_id,
                DoNotCallEntry.company_id == self._company_id(user),
            )
        )
        if not entry:
            raise NotFoundError("Do-not-call entry not found")
        await self.db.delete(entry)
        await self.db.commit()

    async def single_call(
        self, data: SingleOutboundCallRequest, user: CurrentUser
    ) -> CampaignResponse:
        if not E164.fullmatch(data.destination_number):
            raise ValidationError("destination_number must be E.164")
        company_id = self._company_id(user)
        blocked = await self.db.scalar(
            select(DoNotCallEntry.id).where(
                DoNotCallEntry.company_id == company_id,
                DoNotCallEntry.phone_number == data.destination_number,
            )
        )
        if blocked:
            raise ValidationError("Destination is on the do-not-call list")
        campaign_data = CampaignCreate(
            name=f"Ad-hoc call {datetime.now(timezone.utc):%Y-%m-%d %H:%M}",
            campaign_type=data.campaign_type,
            phone_number_id=data.phone_number_id,
            agent_id=data.agent_id,
            message_text=data.message_text,
            voice=data.voice,
            language=data.language,
            timezone="UTC",
            calling_window_start=time(0, 0),
            calling_window_end=time(23, 59, 59),
            max_attempts=1,
        )
        response = await self.create(campaign_data, user)
        campaign = await self._get(response.id, user)
        self.db.add(
            OutboundRecipient(
                campaign_id=campaign.id,
                company_id=campaign.company_id,
                phone_number=data.destination_number,
            )
        )
        campaign.status = CampaignStatus.READY
        await self.db.commit()
        if campaign.campaign_type != CampaignType.AI_CONVERSATION:
            await self.generate_audio(campaign.id, user)
        return await self.start(campaign.id, user)

    async def apply_event(self, data: OutboundEventRequest) -> None:
        row = (
            await self.db.execute(
                select(OutboundAttempt, OutboundRecipient, OutboundCampaign)
                .join(
                    OutboundRecipient,
                    OutboundRecipient.id == OutboundAttempt.recipient_id,
                )
                .join(
                    OutboundCampaign, OutboundCampaign.id == OutboundAttempt.campaign_id
                )
                .where(OutboundAttempt.id == data.attempt_id)
            )
        ).one_or_none()
        if not row:
            raise NotFoundError("Outbound attempt not found")
        attempt, recipient, campaign = row
        now = data.timestamp or datetime.now(timezone.utc)
        attempt.status = data.status
        attempt.provider_call_id = data.provider_call_id or attempt.provider_call_id
        if data.status == RecipientStatus.ANSWERED:
            attempt.answered_at = now
            recipient.status = RecipientStatus.ANSWERED
        elif data.status in {
            RecipientStatus.COMPLETED,
            RecipientStatus.BUSY,
            RecipientStatus.NO_ANSWER,
            RecipientStatus.FAILED,
            RecipientStatus.CANCELLED,
            RecipientStatus.DO_NOT_CALL,
        }:
            attempt.ended_at = now
            attempt.failure_reason = data.reason
            retryable = data.status in {
                RecipientStatus.BUSY,
                RecipientStatus.NO_ANSWER,
                RecipientStatus.FAILED,
            }
            if (
                retryable
                and recipient.attempts_count < campaign.max_attempts
                and campaign.status == CampaignStatus.RUNNING
            ):
                recipient.status = RecipientStatus.PENDING
                recipient.next_attempt_at = now + timedelta(
                    minutes=campaign.retry_delay_minutes
                )
                recipient.last_error = data.reason
            else:
                recipient.status = data.status
                recipient.last_error = data.reason
                if data.status == RecipientStatus.DO_NOT_CALL:
                    exists = await self.db.scalar(
                        select(DoNotCallEntry.id).where(
                            DoNotCallEntry.company_id == recipient.company_id,
                            DoNotCallEntry.phone_number == recipient.phone_number,
                        )
                    )
                    if not exists:
                        self.db.add(
                            DoNotCallEntry(
                                company_id=recipient.company_id,
                                phone_number=recipient.phone_number,
                                reason=data.reason or "Recipient opted out",
                            )
                        )
        from app.modules.calls.models import Call, CallStatus

        if attempt.call_id:
            call = await self.db.get(Call, attempt.call_id)
            if call:
                if data.status == RecipientStatus.ANSWERED:
                    call.status = CallStatus.IN_PROGRESS
                    call.answered_at = now
                elif data.status in {
                    RecipientStatus.COMPLETED,
                    RecipientStatus.DO_NOT_CALL,
                    RecipientStatus.CANCELLED,
                }:
                    call.status = CallStatus.COMPLETED
                    call.ended_at = now
                    call.duration_seconds = data.duration_seconds
                elif data.status in {
                    RecipientStatus.BUSY,
                    RecipientStatus.NO_ANSWER,
                    RecipientStatus.FAILED,
                }:
                    call.status = (
                        CallStatus.FAILED
                        if data.status == RecipientStatus.FAILED
                        else CallStatus.MISSED
                    )
                    call.ended_at = now
        await self.db.commit()
        remaining = int(
            await self.db.scalar(
                select(func.count(OutboundRecipient.id)).where(
                    OutboundRecipient.campaign_id == campaign.id,
                    OutboundRecipient.status.in_(
                        [
                            RecipientStatus.PENDING,
                            RecipientStatus.QUEUED,
                            RecipientStatus.DIALING,
                            RecipientStatus.RINGING,
                            RecipientStatus.ANSWERED,
                        ]
                    ),
                )
            )
            or 0
        )
        if remaining == 0 and campaign.status == CampaignStatus.RUNNING:
            campaign.status = CampaignStatus.COMPLETED
            await self.db.commit()
        elif remaining and campaign.status == CampaignStatus.RUNNING:
            from app.workers.outbound_tasks import dispatch_campaign

            dispatch_campaign.delay(str(campaign.id))

    async def test_call(
        self, campaign_id: uuid.UUID, data: CampaignTestCallRequest, user: CurrentUser
    ) -> CampaignResponse:
        source = await self._get(campaign_id, user)
        if not E164.fullmatch(data.destination_number):
            raise ValidationError("destination_number must be E.164")
        if (
            source.campaign_type != CampaignType.AI_CONVERSATION
            and not source.audio_media_id
        ):
            raise ValidationError("Generate campaign audio before making a test call")
        if await self.db.scalar(
            select(DoNotCallEntry.id).where(
                DoNotCallEntry.company_id == source.company_id,
                DoNotCallEntry.phone_number == data.destination_number,
            )
        ):
            raise ValidationError("Destination is on the do-not-call list")
        response = await self.create(
            CampaignCreate(
                name=f"Test: {source.name}",
                campaign_type=source.campaign_type,
                phone_number_id=source.phone_number_id,
                agent_id=source.agent_id,
                message_text=source.message_text,
                voice=source.voice,
                language=source.language,
                timezone="UTC",
                calling_window_start=time(0, 0),
                calling_window_end=time(23, 59, 59),
                ring_timeout_seconds=source.ring_timeout_seconds,
                keypad_actions=source.keypad_actions,
                max_attempts=1,
            ),
            user,
        )
        test_campaign = await self._get(response.id, user)
        if source.campaign_type != CampaignType.AI_CONVERSATION:
            test_campaign.audio_media_id = source.audio_media_id
            test_campaign.audio_storage_key = source.audio_storage_key
        self.db.add(
            OutboundRecipient(
                campaign_id=test_campaign.id,
                company_id=test_campaign.company_id,
                phone_number=data.destination_number,
                external_id="test-call",
            )
        )
        test_campaign.status = CampaignStatus.READY
        await self.db.commit()
        return await self.start(test_campaign.id, user)
